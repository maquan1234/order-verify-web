# -*- coding: utf-8 -*-
"""
审单核对模块 - 从 verify_orders_v2.py 移植
支持动态加载货品信息、文件上传核对、结果导出
"""
import re
import json
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict, OrderedDict
import tempfile

# ========== 货品信息管理 ==========

def load_products(json_path):
    """从JSON文件加载货品信息"""
    if os.path.exists(json_path):
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_products(json_path, products):
    """保存货品信息到JSON文件"""
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)

def safe_get(row, idx, default=''):
    """安全获取tuple/list中的元素，越界时返回默认值"""
    if not isinstance(row, (tuple, list)):
        return default
    if idx < len(row):
        val = row[idx]
        return val if val is not None else default
    return default

def init_products_from_xlsx(xlsx_path, json_path):
    """从货品信息表xlsx初始化products.json"""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb['Sheet1']
    products = []
    pid = 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = safe_get(row, 0)
        code = safe_get(row, 1)
        barcode = safe_get(row, 2)
        common = safe_get(row, 3)
        if barcode is None or str(barcode).strip() == '':
            continue
        products.append({
            'id': pid,
            'name': str(name) if name else '',
            'code': str(code) if code else '',
            'barcode': str(barcode).strip(),
            'common': str(common) if common else '',
        })
        pid += 1
    wb.close()
    save_products(json_path, products)
    return products

def _build_single_product_info(prod):
    """构建单个产品的info字典（公共逻辑）"""
    name_str = prod.get('name', '')
    common_str = prod.get('common', '')
    barcode_str = str(prod.get('barcode', '')).strip()

    box_count = 1
    base_barcode = barcode_str
    box_count_from_suffix = False
    for suffix, count in [('-9', 9), ('-6', 6), ('-3', 3), ('-2', 2), ('-1', 1)]:
        if barcode_str.endswith(suffix):
            box_count = count
            base_barcode = barcode_str[:-len(suffix)]
            box_count_from_suffix = True
            break

    inner_count = box_count if box_count_from_suffix else 1
    if not box_count_from_suffix and name_str:
        m = re.search(r'[\*×](\d+)(?![\.\d])', name_str)
        if m:
            inner_count = int(m.group(1))
        else:
            m = re.search(r'(\d+)\s*(支装|只装)', name_str)
            if m:
                inner_count = int(m.group(1))

    common_names = [c.strip() for c in common_str.split('/')] if common_str else []

    package_type = '瓶装'
    bags_per_pack = 1
    if '袋装' in name_str or re.search(r'\d+袋', name_str):
        package_type = '袋装'
        m = re.search(r'(\d+)\s*袋', name_str)
        if m:
            bags_per_pack = int(m.group(1))

    return {
        'name': name_str,
        'code': str(prod.get('code', '')).strip(),
        'barcode': barcode_str,
        'common_names': common_names,
        'box_count': box_count,
        'box_count_from_suffix': box_count_from_suffix,
        'inner_count': inner_count,
        'base_barcode': base_barcode,
        'package_type': package_type,
        'bags_per_pack': bags_per_pack,
    }

def build_barcode_to_info(products):
    """从货品列表构建 barcode_to_info 映射"""
    barcode_to_info = {}
    for prod in products:
        barcode_str = str(prod.get('barcode', '')).strip()
        if not barcode_str:
            continue
        barcode_to_info[barcode_str] = _build_single_product_info(prod)
    return barcode_to_info

def build_code_to_info(products):
    """从货品列表构建 code_to_info 映射（当无条码时用货品编号匹配）"""
    code_to_info = {}
    for prod in products:
        code_str = str(prod.get('code', '')).strip()
        if not code_str:
            continue
        code_to_info[code_str] = _build_single_product_info(prod)
    return code_to_info

# ========== 产品分类映射 ==========

def get_product_category(barcode, barcode_to_info, code=None, code_to_info=None):
    info = barcode_to_info.get(barcode)
    if not info and code and code_to_info:
        info = code_to_info.get(code)
    if not info:
        return '未知'
    name = info['name']
    common = '/'.join(info['common_names'])

    if 'EGCG饮3.0' in name or '升级肽3.0' in common:
        return '升级肽3.0'
    if 'EGCG饮2.0' in name and '枇杷' in name:
        return '升级肽2.0枇杷味'
    if 'EGCG饮2.0' in name or '升级肽2.0' in common:
        return '升级肽2.0'
    if 'EGCG饮1.0' in name or '升级肽1.0' in common:
        return '升级肽1.0'
    if '胶原蛋白肽维C饮品2.0' in name or '维C饮2.0' in name:
        return '经典肽2.0'
    if '胶原蛋白肽饮3.0' in name or '经典肽3.0' in common:
        return '经典肽3.0'
    if '维C饮1.0' in name or '经典肽1.0' in common:
        return '经典肽1.0'
    if '高端肽' in name or '高端肽' in common:
        return '高端肽'
    if 'PQQ' in name or 'PQQ' in common:
        return 'PQQ饮'
    if '虾青素白松茸' in name or '虾青素Pro' in common:
        return '虾青素Pro'
    if '虾青素凝胶糖果-6粒' in name or '虾青素6粒' in common or '虾青素试吃' in common:
        return '虾青素6粒'
    if '虾青素凝胶糖果' in name and '6粒' not in name:
        return '虾青素1.0'
    if '超级零' in name and '虾青素' in name:
        return '虾青素饮'
    if '富铁软糖' in name or '软糖' in common:
        if '降糖' in name or '减糖' in common or '降糖版' in common:
            if '6粒' in name:
                return '富铁软糖降糖版6粒'
            return '富铁软糖降糖版'
        if '常糖' in common:
            return '富铁软糖常糖版'
        if '7mg' in common:
            return '富铁软糖7mg'
        return '富铁软糖'
    if '白芸豆阿拉伯糖' in name or '白芸豆' in common:
        return '白芸豆粉剂'
    if '亢唐' in name or '抗糖' in name or '抗糖' in common or '亢糖' in common:
        return '抗糖'
    if '白密码' in name or '白番茄' in name or '白密码' in common or '白番茄饮' in common or '美白饮' in common:
        return '白密码'
    if '美白淡斑精华' in name or '美白精华' in common or '淡斑精华' in common:
        if '5只' in common or '5支装' in name:
            return '美白精华5支装'
        return '美白精华'
    if '洁面' in name or '洁面' in common or '洗面奶' in common:
        return '洁面'
    if '修复' in name or '修复贴' in common:
        return '修复贴'
    if '复合益生菌固体饮料3g*7' in name or ('益生菌' in common and '5只' not in common and '冻干粉' not in common):
        return '益生菌'
    if '复合益生菌-5支' in name or '益生菌5只' in common:
        return '益生菌5支装'
    if '益生菌冻干粉' in name or '益生菌冻干粉' in common:
        if '3袋' in common or '3g*3' in name:
            return '益生菌冻干粉3袋'
        return '益生菌冻干粉'
    if '红参铁' in name or '元气饮' in common or '富铁饮' in common or '红参铁' in common:
        return '元气饮'
    if '褪黑素' in name or '褪黑素' in common:
        return '褪黑素'
    if '钙维生素D3' in name or '女钙' in common or '成人钙' in common or '女性钙' in common:
        if '2只' in common:
            return '女钙2支装'
        return '女钙'
    if '双钙营养包' in name or '儿童钙' in common:
        if '单只' in common:
            return '儿童钙单支'
        return '儿童钙'
    if '小博士钙铁锌' in name or '高高钙' in common:
        if '2只' in common:
            return '高高钙2支装'
        return '高高钙'
    if '多维片' in name or '多维' in common:
        return '多维'
    if '牛初乳' in name or '牛初乳' in common:
        return '牛初乳'
    if '磷虾油' in name or '磷虾油' in common:
        return '磷虾油'
    if '水润面膜' in name or '水润面膜' in common:
        return '水润面膜'
    if '开瓶器' in name or '开瓶器' in common:
        if '30ml' in name or '30ml' in common or '紫色' in common or '经典肽开瓶器' in common:
            return '30ml开瓶器'
        if '50ml' in name or '50ml' in common or '粉色' in common or '升级肽开瓶器' in common:
            return '50ml开瓶器'
        if '大开瓶器' in common:
            return '大开瓶器'
        return '开瓶器'
    if '手提袋' in name or '手提袋' in common or '礼品袋' in common:
        return '手提袋'
    if '书包' in name or '书包' in common:
        return '书包'
    if '身高贴' in name or '身高贴' in common:
        return '身高贴'
    if '梳子' in name or '梳子' in common or '按摩梳' in name:
        return '梳子'
    if '行李箱' in name or '行李箱' in common:
        return '行李箱'
    if '虾青素盲盒' in name or '虾青素盲盒' in common:
        return '虾青素盲盒'
    if '小熊' in name or '小熊盲盒' in common:
        return '小熊盲盒'
    if '相伴礼盒' in name or '相伴礼盒' in common:
        return '相伴礼盒'
    return f'其他:{name}'

# ========== 备注解析器 ==========

CN_NUM_MAP = {'一': 1, '二': 2, '两': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '十': 10}

def parse_cn_number(text):
    if text in CN_NUM_MAP:
        return CN_NUM_MAP[text]
    if text.startswith('十') and len(text) == 2 and text[1] in CN_NUM_MAP:
        return 10 + CN_NUM_MAP[text[1]]
    return None

PRODUCT_KEYWORDS = [
    ('50ml开瓶器', ['50ml开瓶器', '50毫升开瓶器', '粉色开瓶器', '升级肽开瓶器']),
    ('30ml开瓶器', ['30ml开瓶器', '30毫升开瓶器', '紫色开瓶器', '经典肽开瓶器']),
    ('大开瓶器', ['大开瓶器']),
    ('开瓶器', ['开瓶器']),
    ('升级肽3.0', ['升级肽3.0', '升级肽3', 'egcg3.0', 'egcg3', '升级3.0', '升级3']),
    ('升级肽2.0', ['升级肽2.0', '升级肽2', 'egcg2.0', 'egcg2']),
    ('升级肽1.0', ['升级肽1.0', '升级肽1', 'egcg1.0', 'egcg1']),
    ('升级肽', ['升级肽', 'egcg']),
    ('经典肽3.0', ['经典肽3.0', '经典肽3', '3.0经典肽', '3.0版本经典肽', '经典肽3.0版本']),
    ('经典肽2.0', ['经典肽2.0', '经典肽2']),
    ('经典肽1.0', ['经典肽1.0', '经典肽1']),
    ('经典肽', ['经典肽']),
    ('高端肽', ['高端肽']),
    ('PQQ饮', ['pqq']),
    ('虾青素Pro', ['虾青素pro', '虾青素por', '虾青素pro版', '虾青pro', '虾青素 pro', '虾青素pro']),
    ('虾青素饮', ['虾青素饮']),
    ('虾青素6粒', ['6粒装虾青素', '6粒虾青素', '虾青素6粒', '6粒装虾青', '6粒虾青']),
    ('虾青素盲盒', ['虾青素盲盒']),
    ('虾青素', ['虾青素', '虾青']),
    ('富铁软糖降糖版6粒', ['6粒装富铁软糖', '6粒富铁软糖', '6粒装软糖', '6粒软糖', '富铁软糖6粒', '6粒富铁']),
    ('富铁软糖降糖版', ['减糖版富铁软糖', '降糖版富铁软糖', '减甜版富铁软糖', '减糖版软糖', '降糖版软糖', '减甜版软糖', '补铁软糖减甜版', '补铁软糖减糖版']),
    ('富铁软糖常糖版', ['常糖版富铁软糖', '常糖版软糖']),
    ('富铁软糖7mg', ['7mg富铁软糖', '富铁软糖7mg']),
    ('富铁软糖', ['富铁软糖', '补铁软糖', '软糖']),
    ('白芸豆粉剂', ['白芸豆粉剂', '白芸豆片', '白芸豆粉', '白芸豆']),
    ('抗糖', ['抗糖', '亢糖', '亢唐', '白芸豆抗糖粉剂', '抗糖粉剂']),
    ('白密码', ['白密码', '白番茄饮', '美白饮', '白番茄']),
    ('美白精华', ['美白精华', '淡斑精华', '美白淡斑精华', '精华']),
    ('洁面', ['洁面', '洗面奶', '氨基酸洁面']),
    ('修复贴', ['修复贴', '修复']),
    ('益生菌5支装', ['益生菌5支', '益生菌5只']),
    ('益生菌冻干粉', ['益生菌冻干粉']),
    ('益生菌', ['益生菌']),
    ('元气饮', ['元气饮', '富参饮', '红参饮', '富铁饮', '红参铁']),
    ('褪黑素', ['褪黑素']),
    ('女钙', ['液体钙', '女钙', '成人钙', '女性钙']),
    ('儿童钙', ['儿童钙']),
    ('高高钙', ['高高钙']),
    ('多维', ['多维']),
    ('牛初乳', ['牛初乳']),
    ('磷虾油', ['磷虾油']),
    ('水润面膜', ['水润面膜', '面膜']),
    ('手提袋', ['手提袋', '礼品袋']),
    ('书包', ['书包']),
    ('身高贴', ['身高贴']),
    ('梳子', ['梳子', '按摩梳']),
    ('行李箱', ['行李箱']),
    ('小熊盲盒', ['小熊盲盒', '小熊']),
    ('相伴礼盒', ['相伴礼盒']),
]

def find_product_in_text(text):
    text_lower = text.lower()
    for category, keywords in PRODUCT_KEYWORDS:
        for kw in keywords:
            kw_lower = kw.lower()
            idx = text_lower.find(kw_lower)
            if idx >= 0:
                return category, idx, len(kw)
    if '便携装' in text_lower and '软糖' not in text_lower and '虾青素' not in text_lower and '虾青' not in text_lower:
        idx = text_lower.find('便携装')
        return '升级肽', idx, 3
    return None, -1, 0

def extract_spec_from_text(text):
    spec = ''
    m = re.search(r'(\d+)\s*(?:ml|毫升)', text, re.IGNORECASE)
    if m:
        spec = m.group(1) + 'ml'
        text = re.sub(r'\d+\s*(?:ml|毫升)', '', text, flags=re.IGNORECASE).strip()
    m = re.search(r'(\d+)\s*粒', text)
    if m:
        spec = m.group(1) + '粒'
        text = re.sub(r'\d+\s*粒', '', text).strip()
    if not spec:
        m = re.search(r'(一|二|两|三|四|五|六|七|八|九|十)\s*粒', text)
        if m:
            cn_val = parse_cn_number(m.group(1))
            if cn_val:
                spec = str(cn_val) + '粒'
                text = re.sub(r'(一|二|两|三|四|五|六|七|八|九|十)\s*粒', '', text).strip()
    if '便携装' in text:
        m = re.search(r'(\d+)\s*袋\s*便携装', text)
        if m:
            spec = m.group(1) + '袋'
            text = re.sub(r'\d+\s*袋\s*便携装', '便携装', text).strip()
    if not spec:
        m = re.search(r'(\d+)\s*袋装', text)
        if m:
            spec = m.group(1) + '袋'
            text = re.sub(r'\d+\s*袋装', '', text).strip()
    if not spec:
        m = re.search(r'(一|二|两|三|四|五|六|七|八|九|十)\s*袋装', text)
        if m:
            cn_val = parse_cn_number(m.group(1))
            if cn_val:
                spec = str(cn_val) + '袋'
                text = re.sub(r'(一|二|两|三|四|五|六|七|八|九|十)\s*袋装', '', text).strip()
    return spec, text

def extract_qty_and_unit(text):
    m = re.search(r'(\d+)\s*(盒|瓶|个|袋|支|只|箱|份)', text)
    if m:
        qty = int(m.group(1))
        unit = m.group(2)
        remaining = text[:m.start()] + text[m.end():]
        return qty, unit, remaining
    m = re.search(r'(一|二|两|三|四|五|六|七|八|九|十)\s*(盒|瓶|个|袋|支|只|箱|份)', text)
    if m:
        qty = parse_cn_number(m.group(1))
        if qty:
            unit = m.group(2)
            remaining = text[:m.start()] + text[m.end():]
            return qty, unit, remaining
    m = re.search(r'[*×x]\s*(\d+)', text)
    if m:
        qty = int(m.group(1))
        remaining = text[:m.start()] + text[m.end():]
        return qty, '', remaining
    for m in re.finditer(r'(?<!\d)(\d{1,2})(?!\d)(?!\.\d)(?!\s*(?:ml|毫升|粒))', text):
        num = int(m.group(1))
        if num > 0 and num < 100:
            remaining = text[:m.start()] + text[m.end():]
            return num, '', remaining
    return None, '', text

def adjust_category_with_spec(category, spec):
    if spec == '6粒':
        if category == '虾青素':
            return '虾青素6粒'
        if category == '富铁软糖':
            return '富铁软糖降糖版6粒'
    return category

def parse_segment(seg):
    seg = seg.strip()
    if not seg or len(seg) < 1:
        return None
    original_seg = seg
    seg = re.sub(r'^(备注[:：]?\s*|私域[-—:]?\s*)', '', seg)
    seg = re.sub(r'^(发货[：:]?|首发[：:]?|首单[：:]?|本次发出[：:]?|首次发货[：:]?|此次发出[：:]?|全部发[：:]?|发出[：:]?|实发[：:]?|补发[：:]?|发)\s*', '', seg)
    seg = re.sub(r'^(首发|首单|本次发出|首次发货|此次发出|全部发|发出|实发|补发)\s*', '', seg)
    seg = re.sub(r'^(加送|赠送|随单发|额外|补发|兑换|多送|加发|送|赠品发|礼包[：:]?)\s*', '', seg)
    seg = re.sub(r'^扣除积分\s*', '', seg)
    seg = re.sub(r'^\+\s*', '', seg)
    seg = re.sub(r'^\d{2,4}[.\-/]\d{1,2}[.\-/]\d{1,2}\s*', '', seg)
    seg = re.sub(r'^\d{6,8}\s*', '', seg)
    seg = re.sub(r'^\d+\.\d+号?\s*', '', seg)
    seg = re.sub(r'【[^】]*】', '', seg)
    seg = re.sub(r'【[^】]*】', '', seg)
    seg = re.sub(r'^(喝|的|版本)\s*', '', seg)
    seg = seg.strip()
    if not seg:
        return None

    category, prod_idx, prod_len = find_product_in_text(seg)
    if category is None:
        return None
    spec, seg_cleaned = extract_spec_from_text(seg)
    category, prod_idx, prod_len = find_product_in_text(seg_cleaned)
    qty, unit, seg_after_qty = extract_qty_and_unit(seg_cleaned)
    category = adjust_category_with_spec(category, spec)
    if qty is None:
        qty = 1
    is_extra = bool(re.search(r'(加送|赠送|随单发|额外|补发|兑换|送|多送|加发|^\+\d|扣除积分|赠品发|礼包)', original_seg))
    return {
        'product_text': seg,
        'category': category,
        'quantity': qty,
        'unit': unit,
        'spec': spec,
        'is_extra': is_extra,
    }

def try_split_by_spaces(seg):
    count = 0
    temp = seg
    positions = []
    for category, keywords in PRODUCT_KEYWORDS:
        temp_lower = temp.lower()
        for kw in keywords:
            kw_lower = kw.lower()
            idx = temp_lower.find(kw_lower)
            if idx >= 0:
                count += 1
                positions.append((idx, idx + len(kw)))
                break
    if count <= 1:
        return [seg]
    parts = re.split(r'\s{2,}', seg)
    if len(parts) > 1:
        return parts
    result = [seg]
    split_points = []
    for m in re.finditer(r'(\d+)\s*(盒|瓶|个|袋|支|只|箱|份)\s+', seg):
        after = seg[m.end():]
        cat_after, _, _ = find_product_in_text(after[:20])
        if not cat_after:
            continue
        before = seg[:m.start()]
        cat_before, _, _ = find_product_in_text(before)
        if not cat_before:
            continue
        split_points.append(m.end())
    if split_points:
        result = []
        last = 0
        for sp in split_points:
            result.append(seg[last:sp].strip())
            last = sp
        result.append(seg[last:].strip())
        return [r for r in result if r]
    split_points = []
    for m in re.finditer(r'(\d+)(盒|瓶|个|袋|支|只|箱|份)?', seg):
        if m.start() == 0:
            continue
        before = seg[:m.start()]
        cat_before, _, _ = find_product_in_text(before)
        if not cat_before:
            continue
        after = seg[m.end():]
        if not after:
            continue
        cat_after, _, _ = find_product_in_text(after[:20])
        if not cat_after:
            continue
        remaining = seg[m.end():m.end()+2]
        if remaining and remaining[0] in '粒毫':
            continue
        if m.end() < len(seg) and seg[m.end():m.end()+1] == '.':
            continue
        if m.group(2) == '袋' and m.end() < len(seg) and seg[m.end():m.end()+1] == '装':
            continue
        split_points.append(m.start())
    if split_points:
        result = []
        last = 0
        for sp in split_points:
            result.append(seg[last:sp].strip())
            last = sp
        result.append(seg[last:].strip())
        return [r for r in result if r]
    return [seg]

def parse_remark(remark):
    if not remark or remark.strip() == '':
        return []
    text = remark.strip()
    logistics_patterns = [
        r'发顺丰|发中通|发圆通|发韵达|发EMS|发邮政|发京东|发德邦|发申通|发极兔|发菜鸟',
        r'不发圆通|不发中通|不发韵达|不发顺丰|不发货|不要发',
        r'送货上门|送货到家|放门口|放驿站|放快递柜|本人签收',
        r'[\u4e00-\u9fff\w]{1,10}发(?:最新|新鲜)日期',
        r'[\u4e00-\u9fff\w]{1,10}发\d{1,2}月(?:份)?(?:的)?日期',
        r'发最新日期|发新鲜日期|发\d{1,2}月(?:份)?(?:的)?日期|发\d{4}年\d{1,2}月',
        r'拦截|拒收|退回|退款|退差价|补差价|差价',
        r'直播间抽奖半价|直播间免单|抽奖半价|免单|中奖',
        r'催发货|催审核|加急|尽快|辛苦|谢谢|拜托',
        r'最新日期|新鲜日期',
    ]
    for pattern in logistics_patterns:
        text = re.sub(pattern, ' ', text)
    cleaned_for_check = re.sub(r'[a-zA-Z0-9_\-\s\|\n\r]', '', text)
    if not cleaned_for_check or len(cleaned_for_check) < 2:
        return []

    def line_has_product_request(line):
        line_stripped = line.strip()
        if not line_stripped:
            return False
        if re.match(r'^(追加|补发|加发|加送|赠送|随单发|兑换)', line_stripped):
            return True
        if re.search(r'\d+\s*(盒|瓶|袋|支|只|粒|个|罐|桶|套|盒装|瓶装|袋装|支装)', line_stripped):
            return True
        for category, keywords in PRODUCT_KEYWORDS:
            for kw in keywords:
                if kw in line_stripped:
                    return True
        return False

    lines = text.split('\n')
    product_lines = [line for line in lines if line_has_product_request(line)]
    if not product_lines:
        return []
    text = '\n'.join(product_lines)

    lines = text.split('\n')
    deduped_lines = []
    seen_blocks = set()
    for line in lines:
        line_stripped = line.strip()
        if line_stripped and line_stripped not in seen_blocks:
            is_dup = False
            for seen in seen_blocks:
                if line_stripped in seen or seen in line_stripped:
                    is_dup = True
                    break
            if not is_dup:
                seen_blocks.add(line_stripped)
                deduped_lines.append(line)
        elif not line_stripped:
            deduped_lines.append(line)
    text = '\n'.join(deduped_lines)

    gift_items = []
    gift_pattern = re.compile(r'[（(]([^）)]+)[)）]')
    for m in gift_pattern.finditer(text):
        content = m.group(1).strip()
        if '+' in content or '＋' in content:
            sub_segments = re.split(r'[+＋]\s*', content)
        else:
            content_merged = content.replace('，', '').replace(',', '')
            sub_segments = re.split(r'\s+', content_merged)
        for seg in sub_segments:
            seg = seg.strip()
            if not seg or len(seg) < 2:
                continue
            item = parse_segment(seg)
            if item:
                gift_items.append(item)

    text_no_parens = gift_pattern.sub(' ', text)
    text_no_parens = re.sub(r'^\s*\+', '加送', text_no_parens)
    segments = re.split(r'[+＋，,。、\n；;]\s*', text_no_parens)

    merged_segments = []
    i = 0
    while i < len(segments):
        seg = segments[i].strip()
        if not seg:
            i += 1
            continue
        cat, _, _ = find_product_in_text(seg)
        if not cat and len(seg) <= 12 and '月' not in seg and re.search(r'(\d+|[一二两三四五六七八九十]+)\s*(?:ml|毫升|粒|盒|瓶|袋|支|只|个|箱)?', seg):
            if i + 1 < len(segments):
                next_seg = segments[i + 1].strip()
                next_cat, next_idx, next_len = find_product_in_text(next_seg)
                if next_cat:
                    product_kw = next_seg[next_idx:next_idx + next_len]
                    merged_segments.append(seg + product_kw)
                    i += 2
                    continue
        merged_segments.append(seg)
        i += 1
    segments = merged_segments

    items = []
    for seg in segments:
        seg = seg.strip()
        if not seg:
            continue
        if re.match(r'^(辛苦|谢谢|发顺丰|顺丰|最新日期|发最新|要求发|送货上门|其余|见小程序|婷婷代拍|永不退款|客户指定|优先安排|发2026|产品比较多|备注一定要|日期|订单编号|成交时间)', seg):
            continue
        if re.match(r'^(发)?顺丰', seg) and len(seg) <= 10:
            continue
        sub_segs = try_split_by_spaces(seg)
        for ss in sub_segs:
            ss = ss.strip()
            if not ss:
                continue
            item = parse_segment(ss)
            if item:
                items.append(item)

    items.extend(gift_items)

    original_text = remark.strip()
    if re.match(r'^(赠品发|赠品[：:]?|随单发)', original_text):
        for item in items:
            item['is_extra'] = True

    return items

# ========== 分类匹配 ==========

def categories_match(remark_cat, actual_cat):
    if remark_cat is None or actual_cat is None:
        return False
    if remark_cat == actual_cat:
        return True
    if remark_cat == '升级肽' and actual_cat.startswith('升级肽'):
        return True
    if remark_cat == '升级肽2.0' and actual_cat in ('升级肽2.0', '升级肽2.0枇杷味'):
        return True
    if remark_cat == '经典肽' and actual_cat.startswith('经典肽'):
        return True
    if remark_cat == '虾青素' and actual_cat in ('虾青素1.0',):
        return True
    if remark_cat == '富铁软糖' and actual_cat in ('富铁软糖常糖版', '富铁软糖降糖版', '富铁软糖7mg', '富铁软糖'):
        return True
    if remark_cat == '美白精华' and actual_cat.startswith('美白精华'):
        return True
    if remark_cat == '益生菌' and actual_cat in ('益生菌', '益生菌5支装', '益生菌冻干粉', '益生菌冻干粉3袋'):
        return True
    if remark_cat == '开瓶器' and actual_cat in ('30ml开瓶器', '50ml开瓶器', '大开瓶器', '开瓶器'):
        return True
    if remark_cat == '钙' and (actual_cat.startswith('女钙') or actual_cat.startswith('儿童钙') or actual_cat.startswith('高高钙')):
        return True
    return False

def get_match_group(category, package_type=None):
    if category.startswith('升级肽'):
        if package_type == '袋装':
            return '升级肽袋装'
        else:
            return '升级肽瓶装'
    if category.startswith('经典肽'):
        return '经典肽'
    if category.startswith('虾青素') and 'Pro' not in category and '6粒' not in category and '饮' not in category and '盲盒' not in category:
        return '虾青素'
    if category.startswith('富铁软糖'):
        return '富铁软糖'
    if category.startswith('美白精华'):
        return '美白精华'
    if category.startswith('益生菌'):
        return '益生菌'
    if category in ('30ml开瓶器', '50ml开瓶器', '大开瓶器', '开瓶器'):
        return '开瓶器'
    if category.startswith('女钙') or category.startswith('儿童钙') or category.startswith('高高钙'):
        return '钙'
    return category

def get_remark_package_type(unit, spec, product_text=''):
    if unit == '袋':
        return '袋装'
    if spec and '袋' in spec:
        return '袋装'
    if product_text and '袋装' in product_text:
        return '袋装'
    if unit in ('盒', '瓶', '箱'):
        return '瓶装'
    return '瓶装'

# ========== 匹配列（sheet0 → sheet1）==========

def _find_col_index(headers, possible_names, after_col=-1, find_last=False):
    """根据表头名称列表找到列索引，返回-1表示未找到。
    after_col: 只搜索该列之后的列（用于避免找到重复列名中的前一个）
    find_last: 为True时返回最后一个匹配（用于行级备注等可能有重复列名的情况）"""
    start = after_col + 1 if after_col >= 0 else 0
    result = -1
    for i in range(start, len(headers)):
        h = headers[i]
        h_str = str(h).strip() if h else ''
        if not h_str:
            continue
        for name in possible_names:
            if name in h_str:
                if find_last:
                    result = i  # 继续搜索，保留最后一个
                else:
                    return i
    return result

def _read_headers(ws):
    """读取工作表第一行作为表头列表"""
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(v).strip() if v else '' for v in row]
        break
    return headers

def match_columns_from_workbook(wb_raw):
    """从原始工作簿匹配网店订单号和备注到sheet1，返回匹配后的行列表。
    自动识别表头列，支持不同列顺序的表格。"""
    ws0 = wb_raw['sheet0']
    ws1_raw = wb_raw['sheet1']

    # 读取sheet0表头，动态识别列
    ws0_headers = _read_headers(ws0)
    order_id_col = _find_col_index(ws0_headers, ['订单编号', '原始订单编号', '平台订单号', '主订单编号'])
    web_order_col = _find_col_index(ws0_headers, ['网店订单号', '平台订单号', '线上订单号', '外部订单号'])
    remark_col = _find_col_index(ws0_headers, ['合并备注', '备注'])

    # 读取sheet1表头，动态识别列
    ws1_headers = _read_headers(ws1_raw)
    s1_order_col = _find_col_index(ws1_headers, ['订单编号', '原始订单编号', '平台订单号', '主订单编号'])
    s1_web_order_col = _find_col_index(ws1_headers, ['网店订单号', '平台订单号', '线上订单号', '外部订单号'])  # 部分格式sheet1自带网店订单号
    barcode_col = _find_col_index(ws1_headers, ['条码'])
    code_col = _find_col_index(ws1_headers, ['货品编号'])
    name_col = _find_col_index(ws1_headers, ['货品名称'])
    tag_col = _find_col_index(ws1_headers, ['货品标记'])
    spec_col = _find_col_index(ws1_headers, ['规格'])
    qty_col = _find_col_index(ws1_headers, ['数量'])
    price_col = _find_col_index(ws1_headers, ['单价'])
    discount_amt_col = _find_col_index(ws1_headers, ['优惠'])
    discount_rate_col = _find_col_index(ws1_headers, ['折扣'])
    amount_col = _find_col_index(ws1_headers, ['金额'])
    line_remark_col = _find_col_index(ws1_headers, ['备注'], find_last=True)
    talent_id_col = _find_col_index(ws1_headers, ['达人ID'])
    talent_name_col = _find_col_index(ws1_headers, ['达人名称'])
    unit_col = _find_col_index(ws1_headers, ['单位'])
    gift_col = _find_col_index(ws1_headers, ['赠品'])

    # Build order_map from sheet0
    order_map = {}
    for row in ws0.iter_rows(min_row=2, max_row=ws0.max_row, values_only=True):
        order_id = str(safe_get(row, order_id_col)).strip() if order_id_col >= 0 else ''
        web_order = str(safe_get(row, web_order_col)).strip() if web_order_col >= 0 else ''
        remark = str(safe_get(row, remark_col)).strip() if remark_col >= 0 else ''
        # 过滤掉 'None' 字符串和空值
        if order_id and order_id != 'None' and order_id not in order_map:
            order_map[order_id] = (web_order if web_order != 'None' else '', remark if remark != 'None' else '')

    # Build matched_rows from sheet1
    matched_rows = []
    for row in ws1_raw.iter_rows(min_row=2, max_row=ws1_raw.max_row, values_only=True):
        order_id = str(safe_get(row, s1_order_col)).strip() if s1_order_col >= 0 else ''
        if order_id == 'None':
            order_id = ''
        if order_id and order_id in order_map:
            web_order, merged_remark = order_map[order_id]
        else:
            web_order, merged_remark = '', ''
        # 如果sheet0没匹配到，尝试用sheet1自带的网店订单号
        if not web_order and s1_web_order_col >= 0:
            s1_web = str(safe_get(row, s1_web_order_col)).strip()
            if s1_web and s1_web != 'None':
                web_order = s1_web

        qty_val = safe_get(row, qty_col) if qty_col >= 0 else ''
        try:
            quantity = int(float(qty_val)) if qty_val not in ('', None, 'None') else 0
        except (ValueError, TypeError):
            quantity = 0

        def _clean(val):
            v = str(val).strip() if val is not None else ''
            return '' if v == 'None' else v

        matched_rows.append({
            'order_id': order_id,
            'web_order_id': web_order,
            'remark': merged_remark,
            'barcode': _clean(safe_get(row, barcode_col)) if barcode_col >= 0 else '',
            'product_code': _clean(safe_get(row, code_col)) if code_col >= 0 else '',
            'product_name': _clean(safe_get(row, name_col)) if name_col >= 0 else '',
            'product_tag': _clean(safe_get(row, tag_col)) if tag_col >= 0 else '',
            'quantity': quantity,
            'unit_price': safe_get(row, price_col, None) if price_col >= 0 else None,
            'amount': safe_get(row, amount_col, None) if amount_col >= 0 else None,
            'spec': _clean(safe_get(row, spec_col)) if spec_col >= 0 else '',
            'talent_id': _clean(safe_get(row, talent_id_col)) if talent_id_col >= 0 else '',
            'talent_name': _clean(safe_get(row, talent_name_col)) if talent_name_col >= 0 else '',
            'line_remark': _clean(safe_get(row, line_remark_col)) if line_remark_col >= 0 else '',
            'unit': _clean(safe_get(row, unit_col)) if unit_col >= 0 else '',
            'discount_amount': safe_get(row, discount_amt_col, None) if discount_amt_col >= 0 else None,
            'discount_rate': safe_get(row, discount_rate_col, None) if discount_rate_col >= 0 else None,
            'is_gift': _clean(safe_get(row, gift_col)) if gift_col >= 0 else '',
        })

    return matched_rows

# ========== 核对主逻辑 ==========

def verify_orders(matched_rows, barcode_to_info, code_to_info=None):
    """核对订单，返回结果列表。支持用条码或货品编号匹配货品信息。"""
    if code_to_info is None:
        code_to_info = {}
    orders = OrderedDict()
    for row in matched_rows:
        web_order_id = row['web_order_id']
        order_id = row.get('order_id', '')
        # 当网店订单号为空时，用订单编号兜底分组，避免所有空web_order_id被合并成一条
        group_key = web_order_id if web_order_id else order_id
        if not group_key:
            continue  # 两者都空则跳过该行
        barcode = row['barcode']
        product_code = row.get('product_code', '')
        if group_key not in orders:
            orders[group_key] = {'remark': row['remark'], 'items': [], 'web_order_id': web_order_id, 'order_id_fallback': not web_order_id}
        # 先用条码查，找不到用货品编号查
        item_info = barcode_to_info.get(barcode)
        if not item_info and product_code:
            item_info = code_to_info.get(product_code, {})
        if not item_info:
            item_info = {}
        category = get_product_category(barcode, barcode_to_info, product_code, code_to_info)
        orders[group_key]['items'].append({
            'order_id': order_id,
            'barcode': barcode,
            'product_code': product_code,
            'product_name': row['product_name'],
            'quantity': row['quantity'],
            'is_gift': row['is_gift'],
            'category': category,
            'box_count': item_info.get('box_count', 1),
            'box_count_from_suffix': item_info.get('box_count_from_suffix', False),
            'inner_count': item_info.get('inner_count', 1),
            'package_type': item_info.get('package_type', '瓶装'),
            'bags_per_pack': item_info.get('bags_per_pack', 1),
        })

    results = []
    for group_key, order_data in orders.items():
        remark = order_data['remark']
        items = order_data['items']
        remark_items = parse_remark(remark)

        actual_by_group = defaultdict(lambda: {
            'box_qty': 0, 'piece_qty': 0, 'has_suffix_box_count': False,
            'details': [], 'is_all_gift': True, 'cats': []
        })
        for item in items:
            cat = item['category']
            pkg_type = item['package_type']
            group = get_match_group(cat, pkg_type)
            if pkg_type == '袋装':
                piece_qty = item['quantity'] * item['bags_per_pack']
                detail_suffix = f"袋×{item['bags_per_pack']}"
            else:
                piece_qty = item['quantity'] * item['inner_count']
                detail_suffix = f"内装{item['inner_count']}"
            actual_by_group[group]['box_qty'] += item['quantity']
            actual_by_group[group]['piece_qty'] += piece_qty
            if item.get('box_count_from_suffix', False):
                actual_by_group[group]['has_suffix_box_count'] = True
            actual_by_group[group]['details'].append(f"{item['product_name']}*{item['quantity']}({detail_suffix})")
            actual_by_group[group]['cats'].append(cat)
            if item['is_gift'] != '是':
                actual_by_group[group]['is_all_gift'] = False

        remark_by_group = defaultdict(lambda: {
            'box_qty': 0, 'piece_qty': 0,
            'items': [], 'categories': set()
        })
        for ri in remark_items:
            pkg_type = get_remark_package_type(ri['unit'], ri['spec'], ri.get('product_text', ''))
            group = get_match_group(ri['category'], pkg_type)
            qty = ri['quantity']
            if pkg_type == '袋装' and ri['spec'] and '袋' in ri['spec']:
                m = re.match(r'(\d+)袋', ri['spec'])
                if m and ri['unit'] in ('', '袋'):
                    qty = qty * int(m.group(1))
            if ri['unit'] in ('盒', '箱'):
                remark_by_group[group]['box_qty'] += qty
            else:
                remark_by_group[group]['piece_qty'] += qty
            remark_by_group[group]['items'].append(ri)
            remark_by_group[group]['categories'].add(ri['category'])

        verification_details = []
        is_correct = True
        issues = []
        matched_groups = set()

        for group, remark_data in remark_by_group.items():
            if group == '未知':
                for ri in remark_data['items']:
                    verification_details.append({
                        'remark_item': f"{ri['quantity']}{ri['unit']}{ri['product_text']}",
                        'actual_matched': '',
                        'actual_qty': '',
                        'remark_qty': ri['quantity'],
                        'status': '⚠️备注中有无法识别的货品',
                    })
                continue

            actual_data = actual_by_group.get(group)
            if actual_data:
                matched_groups.add(group)
                remark_desc = ' + '.join([f"{ri['quantity']}{ri['unit']}{ri['product_text']}" for ri in remark_data['items']])
                all_extra = all(ri.get('is_extra', False) for ri in remark_data['items'])
                has_suffix = actual_data.get('has_suffix_box_count', False)
                if has_suffix:
                    remark_piece_qty = remark_data['piece_qty'] + remark_data['box_qty']
                    remark_box_qty = 0
                else:
                    remark_piece_qty = remark_data['piece_qty']
                    remark_box_qty = remark_data['box_qty']
                box_match = actual_data['box_qty'] == remark_box_qty
                piece_match = actual_data['piece_qty'] == remark_piece_qty
                compare_parts = []
                if remark_box_qty > 0:
                    compare_parts.append(f"盒数:实际{actual_data['box_qty']}vs备注{remark_box_qty}")
                if remark_piece_qty > 0:
                    compare_parts.append(f"件数:实际{actual_data['piece_qty']}vs备注{remark_piece_qty}")
                compare_desc = ' | '.join(compare_parts)
                if remark_box_qty > 0 and remark_piece_qty > 0:
                    is_match = box_match and piece_match
                elif remark_box_qty > 0:
                    is_match = box_match
                else:
                    is_match = piece_match
                if is_match:
                    verification_details.append({
                        'remark_item': remark_desc,
                        'actual_matched': '; '.join(actual_data['details']),
                        'actual_qty': f"盒{actual_data['box_qty']}/件{actual_data['piece_qty']}",
                        'remark_qty': f"盒{remark_box_qty}/件{remark_piece_qty}",
                        'status': '✅正确',
                    })
                else:
                    if all_extra:
                        extra_ok = True
                        if remark_box_qty > 0 and actual_data['box_qty'] < remark_box_qty:
                            extra_ok = False
                        if remark_piece_qty > 0 and actual_data['piece_qty'] < remark_piece_qty:
                            extra_ok = False
                        if extra_ok:
                            verification_details.append({
                                'remark_item': remark_desc,
                                'actual_matched': '; '.join(actual_data['details']),
                                'actual_qty': f"盒{actual_data['box_qty']}/件{actual_data['piece_qty']}",
                                'remark_qty': f"盒{remark_box_qty}/件{remark_piece_qty}",
                                'status': f'✅匹配(加送/赠送项, {compare_desc})',
                            })
                            continue
                    is_correct = False
                    issues.append(f"备注要求{remark_desc}，{compare_desc}")
                    verification_details.append({
                        'remark_item': remark_desc,
                        'actual_matched': '; '.join(actual_data['details']),
                        'actual_qty': f"盒{actual_data['box_qty']}/件{actual_data['piece_qty']}",
                        'remark_qty': f"盒{remark_box_qty}/件{remark_piece_qty}",
                        'status': f'❌数量不符({compare_desc})',
                    })
            else:
                is_correct = False
                remark_desc = ' + '.join([f"{ri['quantity']}{ri['unit']}{ri['product_text']}" for ri in remark_data['items']])
                issues.append(f"备注要求{remark_desc}但订单中未找到")
                verification_details.append({
                    'remark_item': remark_desc,
                    'actual_matched': '未找到',
                    'actual_qty': 0,
                    'remark_qty': f"盒{remark_data['box_qty']}/件{remark_data['piece_qty']}",
                    'status': '❌缺失',
                })

        all_remark_extra = remark_items and all(ri.get('is_extra', False) for ri in remark_items)
        if remark_items:
            for group, data in actual_by_group.items():
                if group not in matched_groups:
                    if data['is_all_gift']:
                        verification_details.append({
                            'remark_item': '(系统赠品)',
                            'actual_matched': '; '.join(data['details']),
                            'actual_qty': f"盒{data['box_qty']}/件{data['piece_qty']}",
                            'remark_qty': 0,
                            'status': '🎁系统赠品',
                        })
                    elif all_remark_extra:
                        verification_details.append({
                            'remark_item': '(备注仅含赠品要求)',
                            'actual_matched': '; '.join(data['details']),
                            'actual_qty': f"盒{data['box_qty']}/件{data['piece_qty']}",
                            'remark_qty': 0,
                            'status': 'ℹ️正常购买(备注仅含赠品要求)',
                        })
                    else:
                        issues.append(f"订单中有{','.join(data['cats'])}(盒{data['box_qty']}/件{data['piece_qty']})但备注中未提及")
                        verification_details.append({
                            'remark_item': '(备注未提及)',
                            'actual_matched': '; '.join(data['details']),
                            'actual_qty': f"盒{data['box_qty']}/件{data['piece_qty']}",
                            'remark_qty': 0,
                            'status': '⚠️多发货品(备注未提及)',
                        })

        if not remark_items and not issues:
            verification_details.append({
                'remark_item': '(无具体货品要求)',
                'actual_matched': '',
                'actual_qty': '',
                'remark_qty': '',
                'status': 'ℹ️备注无具体货品要求',
            })

        web_order_id = order_data['web_order_id']
        order_ids = ', '.join(sorted(set(item['order_id'] for item in items if item['order_id'])))
        # 如果是用订单编号兜底分组的，把网店订单号显示为空或提示
        display_web_order_id = web_order_id if web_order_id else '(未匹配到网店订单号)'
        results.append({
            'web_order_id': display_web_order_id,
            'order_ids': order_ids,
            'remark': remark,
            'is_correct': is_correct,
            'issues': '; '.join(issues) if issues else '',
            'details': verification_details,
            'items': items,
        })

    return results

# ========== 结果导出 ==========

def export_results_to_xlsx(results, output_path):
    """将核对结果导出为xlsx文件"""
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = '核对结果'

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', start_color='4472C4')
    correct_fill = PatternFill('solid', start_color='C6EFCE')
    error_fill = PatternFill('solid', start_color='FFC7CE')
    warning_fill = PatternFill('solid', start_color='FFEB9C')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    headers = ['网店订单号', '订单编号', '核对结果', '问题说明', '备注内容', '备注要求货品及数量', '实际发货货品及数量', '明细核对']
    for col, h in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    widths = [20, 35, 10, 45, 50, 35, 50, 40]
    for col, w in enumerate(widths, 1):
        ws_out.column_dimensions[chr(64 + col)].width = w

    row_idx = 2
    for r in results:
        remark_items_str = '\n'.join([d['remark_item'] for d in r['details'] if d['remark_item'] and d['remark_item'] not in ['(系统赠品)', '(备注未提及)', '(无具体货品要求)', '(备注仅含赠品要求)']])
        actual_items_str = '\n'.join([f"{d['actual_matched']}" for d in r['details'] if d['actual_matched']])
        detail_str = '\n'.join([f"[{d['status']}] {d['remark_item']}" for d in r['details']])

        has_warnings = bool(r['issues']) and r['is_correct']

        ws_out.cell(row=row_idx, column=1, value=r['web_order_id'])
        ws_out.cell(row=row_idx, column=2, value=r['order_ids'])
        if r['is_correct'] and not has_warnings:
            result_text = '✅正确'
            result_fill = correct_fill
        elif has_warnings:
            result_text = '⚠️有警告'
            result_fill = warning_fill
        else:
            result_text = '❌有误'
            result_fill = error_fill
        result_cell = ws_out.cell(row=row_idx, column=3, value=result_text)
        result_cell.fill = result_fill
        result_cell.alignment = Alignment(horizontal='center', vertical='center')

        issue_cell = ws_out.cell(row=row_idx, column=4, value=r['issues'])
        if r['issues']:
            issue_cell.fill = warning_fill if has_warnings else error_fill
        ws_out.cell(row=row_idx, column=5, value=r['remark'])
        ws_out.cell(row=row_idx, column=6, value=remark_items_str)
        ws_out.cell(row=row_idx, column=7, value=actual_items_str)
        ws_out.cell(row=row_idx, column=8, value=detail_str)

        for col in range(1, 9):
            cell = ws_out.cell(row=row_idx, column=col)
            cell.alignment = wrap_alignment
            cell.border = thin_border
        row_idx += 1

    ws_out.freeze_panes = 'A2'

    ws_summary = wb_out.create_sheet('核对汇总')
    for col, h in enumerate(['统计项', '数量'], 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    total = len(results)
    correct = sum(1 for r in results if r['is_correct'] and not (r['issues'] and r['is_correct']))
    warnings = sum(1 for r in results if r['is_correct'] and r['issues'])
    errors = sum(1 for r in results if not r['is_correct'])

    ws_summary.cell(row=2, column=1, value='总订单数').border = thin_border
    ws_summary.cell(row=2, column=2, value=total).border = thin_border
    ws_summary.cell(row=3, column=1, value='核对正确').border = thin_border
    c = ws_summary.cell(row=3, column=2, value=correct)
    c.fill = correct_fill; c.border = thin_border
    ws_summary.cell(row=4, column=1, value='有警告(多发货品)').border = thin_border
    c = ws_summary.cell(row=4, column=2, value=warnings)
    c.fill = warning_fill; c.border = thin_border
    ws_summary.cell(row=5, column=1, value='核对有误').border = thin_border
    c = ws_summary.cell(row=5, column=2, value=errors)
    c.fill = error_fill; c.border = thin_border
    ws_summary.column_dimensions['A'].width = 18
    ws_summary.column_dimensions['B'].width = 10

    wb_out.save(output_path)
    return {'total': total, 'correct': correct, 'warnings': warnings, 'errors': errors}
