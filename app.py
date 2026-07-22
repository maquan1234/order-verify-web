# -*- coding: utf-8 -*-
"""电商审单核对Web应用 - Flask后端（支持本地桌面打包 / 云部署）"""
import os
import sys
import tempfile
import time
import socket
import webbrowser
from flask import Flask, request, jsonify, send_file, render_template
from verifier import (
    load_products, save_products, init_products_from_xlsx, build_barcode_to_info,
    build_code_to_info, match_columns_from_workbook, verify_orders, export_results_to_xlsx
)
from openpyxl import load_workbook

# ========== 路径解析（兼容 PyInstaller 打包后的 frozen 状态） ==========
def _is_frozen():
    """打包成 exe 后，sys.frozen 为 True，资源在 sys._MEIPASS 内"""
    return getattr(sys, 'frozen', False)

if _is_frozen():
    # 打包后：模板/货品表等资源被解压到 sys._MEIPASS（只读，不可写）
    RESOURCE_DIR = sys._MEIPASS
    EXE_DIR = os.path.dirname(sys.executable)
    # 数据目录放在用户 AppData，保证可写且持久（不同机器各自保存货品数据）
    APPDATA = os.environ.get('APPDATA', EXE_DIR)
    DATA_DIR = os.path.join(APPDATA, 'OrderVerify')
else:
    RESOURCE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXE_DIR = RESOURCE_DIR
    # 云部署 / 开发：DATA_DIR 可用环境变量覆盖（指向持久卷）
    DATA_DIR = os.environ.get('DATA_DIR', RESOURCE_DIR)

os.makedirs(DATA_DIR, exist_ok=True)

PRODUCTS_JSON = os.path.join(DATA_DIR, 'products.json')
# 打包后不再依赖桌面文件；如需指定本地货品表可用环境变量 PRODUCTS_XLSX 覆盖
PRODUCTS_XLSX = os.environ.get('PRODUCTS_XLSX', '')
OUTPUT_DIR = os.path.join(DATA_DIR, 'outputs')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 模板与静态资源：打包后从 RESOURCE_DIR 读取（_MEIPASS 内已包含）
template_folder = os.path.join(RESOURCE_DIR, 'templates')
static_folder = os.path.join(RESOURCE_DIR, 'static') if os.path.isdir(os.path.join(RESOURCE_DIR, 'static')) else None
app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max upload

# ========== 初始化货品信息 ==========
def ensure_products():
    if not os.path.exists(PRODUCTS_JSON):
        # 首次运行：若数据目录为空，但程序自带 products.json（含初始货品），则种子化过去
        seed = os.path.join(RESOURCE_DIR, 'products.json')
        if _is_frozen() and os.path.exists(seed):
            import shutil
            shutil.copyfile(seed, PRODUCTS_JSON)
        elif os.path.exists(PRODUCTS_XLSX):
            init_products_from_xlsx(PRODUCTS_XLSX, PRODUCTS_JSON)
        else:
            save_products(PRODUCTS_JSON, [])

ensure_products()

# ========== 页面路由 ==========
@app.route('/')
def index():
    return render_template('index.html')

# ========== 货品信息API ==========
@app.route('/api/products', methods=['GET'])
def get_products():
    products = load_products(PRODUCTS_JSON)
    return jsonify(products)

@app.route('/api/products', methods=['POST'])
def add_product():
    data = request.json
    products = load_products(PRODUCTS_JSON)
    new_id = max([p['id'] for p in products], default=0) + 1
    data['id'] = new_id
    products.append(data)
    save_products(PRODUCTS_JSON, products)
    return jsonify({'success': True, 'product': data})

@app.route('/api/products/<int:pid>', methods=['PUT'])
def update_product(pid):
    data = request.json
    products = load_products(PRODUCTS_JSON)
    for p in products:
        if p['id'] == pid:
            p['name'] = data.get('name', p['name'])
            p['code'] = data.get('code', p['code'])
            p['barcode'] = data.get('barcode', p['barcode'])
            p['common'] = data.get('common', p['common'])
            save_products(PRODUCTS_JSON, products)
            return jsonify({'success': True, 'product': p})
    return jsonify({'success': False, 'error': '货品不存在'}), 404

@app.route('/api/products/<int:pid>', methods=['DELETE'])
def delete_product(pid):
    products = load_products(PRODUCTS_JSON)
    products = [p for p in products if p['id'] != pid]
    save_products(PRODUCTS_JSON, products)
    return jsonify({'success': True})

@app.route('/api/products/import', methods=['POST'])
def import_products():
    """从xlsx文件导入货品信息（覆盖）"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未上传文件'}), 400
    f = request.files['file']
    tmp_path = os.path.join(tempfile.gettempdir(), f'prod_{int(time.time())}.xlsx')
    f.save(tmp_path)
    try:
        products = init_products_from_xlsx(tmp_path, PRODUCTS_JSON)
        return jsonify({'success': True, 'count': len(products)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

# ========== 审单核对API ==========
@app.route('/api/verify', methods=['POST'])
def verify_upload():
    """上传销售单xlsx，自动核对，返回结果摘要和下载ID"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未上传文件'}), 400
    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': '请上传xlsx文件'}), 400

    tmp_path = os.path.join(tempfile.gettempdir(), f'upload_{int(time.time())}.xlsx')
    f.save(tmp_path)

    try:
        wb_raw = load_workbook(tmp_path, data_only=True)

        # 检查sheet结构，支持不同sheet名称
        sheet_names = wb_raw.sheetnames
        if len(sheet_names) < 2:
            wb_raw.close()
            return jsonify({'success': False, 'error': '文件需至少包含2个工作表（订单信息+货品明细）'}), 400

        # 如果sheet名称不是sheet0/sheet1，自动重命名
        if 'sheet0' not in sheet_names or 'sheet1' not in sheet_names:
            # 尝试根据内容判断哪个是sheet0（订单汇总）哪个是sheet1（货品明细）
            from verifier import _read_headers, _find_col_index
            ws_a = wb_raw[sheet_names[0]]
            ws_b = wb_raw[sheet_names[1]]
            headers_a = _read_headers(ws_a)
            headers_b = _read_headers(ws_b)
            # sheet1有货品编号或条码，sheet0有合并备注或网店订单号
            has_product_b = _find_col_index(headers_b, ['货品编号', '条码']) >= 0
            has_product_a = _find_col_index(headers_a, ['货品编号', '条码']) >= 0
            if has_product_b and not has_product_a:
                wb_raw[sheet_names[0]].title = 'sheet0'
                wb_raw[sheet_names[1]].title = 'sheet1'
            elif has_product_a and not has_product_b:
                wb_raw[sheet_names[0]].title = 'sheet1'
                wb_raw[sheet_names[1]].title = 'sheet0'
            else:
                # 默认第一个是sheet0，第二个是sheet1
                wb_raw[sheet_names[0]].title = 'sheet0'
                wb_raw[sheet_names[1]].title = 'sheet1'

        matched_rows = match_columns_from_workbook(wb_raw)
        wb_raw.close()

        # 诊断：统计未匹配到网店订单号的行数
        unmatched_count = sum(1 for r in matched_rows if not r.get('web_order_id'))
        warn_msg = None
        if unmatched_count > 0:
            total_rows = len(matched_rows)
            warn_msg = f'注意：{unmatched_count}/{total_rows} 行未从sheet0匹配到网店订单号，已按订单编号单独分组。请检查sheet0是否包含对应的订单编号。'

        # 加载货品信息并核对
        products = load_products(PRODUCTS_JSON)
        barcode_to_info = build_barcode_to_info(products)
        code_to_info = build_code_to_info(products)
        results = verify_orders(matched_rows, barcode_to_info, code_to_info)

        # 导出结果
        download_id = f'result_{int(time.time())}.xlsx'
        output_path = os.path.join(OUTPUT_DIR, download_id)
        summary = export_results_to_xlsx(results, output_path)

        # 构建前端展示用的结果摘要
        order_results = []
        for r in results:
            has_warnings = bool(r['issues']) and r['is_correct']
            if r['is_correct'] and not has_warnings:
                status = 'correct'
            elif has_warnings:
                status = 'warning'
            else:
                status = 'error'

            order_results.append({
                'web_order_id': r['web_order_id'],
                'order_ids': r['order_ids'],
                'remark': r['remark'][:500] if r['remark'] else '',
                'status': status,
                'issues': r['issues'],
                'details': r['details'],
            })

        response_data = {
            'success': True,
            'summary': summary,
            'orders': order_results,
            'download_id': download_id,
        }
        if warn_msg:
            response_data['warning'] = warn_msg
        return jsonify(response_data)

    except Exception as e:
        import traceback
        traceback.print_exc()
        err_msg = str(e)
        # 对常见Excel列越界错误给出更友好的提示
        if 'tuple index out of range' in err_msg or 'index out of range' in err_msg:
            err_msg = 'Excel表格列数不足，请确保sheet0和sheet1的列结构与模板一致。常见原因：某行数据被删除或表格格式不正确。'
        return jsonify({'success': False, 'error': err_msg}), 500
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

@app.route('/api/download/<download_id>', methods=['GET'])
def download_result(download_id):
    """下载核对结果文件"""
    # 安全检查：只允许下载outputs目录下的文件
    if '..' in download_id or '/' in download_id or '\\' in download_id:
        return jsonify({'error': '非法文件名'}), 400
    path = os.path.join(OUTPUT_DIR, download_id)
    if not os.path.exists(path):
        return jsonify({'error': '文件不存在'}), 404
    return send_file(path, as_attachment=True, download_name=f'销售单核对结果_{download_id}')

# ========== 企业微信机器人回调（需配置 WECOM_ENABLED=1） ==========
def _wecom_reply(token, chat_id, text):
    try:
        import wecom_bot
        wecom_bot.send_text_to_chat(token, chat_id, text)
    except Exception as e:
        print('[wecom] 回复失败:', e)

def _wecom_process_file(media_id, chat_id):
    """后台线程：下载群里发的 Excel -> 核查 -> 把结果发回群聊。"""
    chat_id = chat_id or ''
    try:
        import wecom_bot
        from verifier import _read_headers, _find_col_index
        corpid = os.environ['WECOM_CORPID']
        secret = os.environ['WECOM_SECRET']
        token = wecom_bot.get_access_token(corpid, secret)

        tmp = os.path.join(tempfile.gettempdir(), f'wecom_{int(time.time())}.xlsx')
        wecom_bot.download_media(token, media_id, tmp)
        try:
            wb_raw = load_workbook(tmp, data_only=True)
            sheet_names = wb_raw.sheetnames
            if len(sheet_names) < 2:
                wb_raw.close()
                _wecom_reply(token, chat_id, '❌ 文件需至少包含 2 个工作表（订单信息 + 货品明细）')
                return
            if 'sheet0' not in sheet_names or 'sheet1' not in sheet_names:
                ws_a = wb_raw[sheet_names[0]]; ws_b = wb_raw[sheet_names[1]]
                ha = _read_headers(ws_a); hb = _read_headers(ws_b)
                hp_b = _find_col_index(hb, ['货品编号', '条码']) >= 0
                hp_a = _find_col_index(ha, ['货品编号', '条码']) >= 0
                if hp_b and not hp_a:
                    wb_raw[sheet_names[0]].title = 'sheet0'; wb_raw[sheet_names[1]].title = 'sheet1'
                elif hp_a and not hp_b:
                    wb_raw[sheet_names[0]].title = 'sheet1'; wb_raw[sheet_names[1]].title = 'sheet0'
                else:
                    wb_raw[sheet_names[0]].title = 'sheet0'; wb_raw[sheet_names[1]].title = 'sheet1'
            matched = match_columns_from_workbook(wb_raw)
            wb_raw.close()

            products = load_products(PRODUCTS_JSON)
            results = verify_orders(matched, build_barcode_to_info(products), build_code_to_info(products))
            download_id = f'result_{int(time.time())}.xlsx'
            out = os.path.join(OUTPUT_DIR, download_id)
            summary = export_results_to_xlsx(results, out)

            base = os.environ.get('PUBLIC_BASE_URL', '').rstrip('/')
            if base:
                link = f'{base}/api/download/{download_id}'
                link_tip = f'结果下载：{link}'
            else:
                link_tip = '（未配置 PUBLIC_BASE_URL，无法生成下载链接；请在环境变量中设置你的公网地址）'
            text = (f'✅ 审单核查完成\n'
                    f'订单总数：{summary.get("total_orders")}\n'
                    f'异常：{summary.get("error_count", 0)}　需关注：{summary.get("warning_count", 0)}\n'
                    f'{link_tip}')
            _wecom_reply(token, chat_id, text)
        finally:
            if os.path.exists(tmp):
                os.remove(tmp)
    except Exception as e:
        import traceback; traceback.print_exc()
        try:
            import wecom_bot
            t2 = wecom_bot.get_access_token(os.environ['WECOM_CORPID'], os.environ['WECOM_SECRET'])
            _wecom_reply(t2, chat_id, f'❌ 处理失败：{e}')
        except Exception:
            pass

@app.route('/wecom/callback', methods=['GET', 'POST'])
def wecom_callback():
    if os.environ.get('WECOM_ENABLED') != '1':
        return 'wecom callback disabled', 403
    try:
        import wecom_bot
        crypto = wecom_bot.WeComCrypto(
            os.environ['WECOM_TOKEN'], os.environ['WECOM_AES_KEY'], os.environ['WECOM_CORPID'])
    except Exception as e:
        return f'wecom config error: {e}', 500
    if request.method == 'GET':
        try:
            plain = crypto.decrypt_echo(
                request.args.get('msg_signature'), request.args.get('timestamp'),
                request.args.get('nonce'), request.args.get('echostr'))
            return plain
        except Exception as e:
            return f'verify failed: {e}', 400
    # POST：先回包，再异步处理，避免企业微信 5s 超时
    try:
        plain = crypto.decrypt_message(
            request.args.get('msg_signature'), request.args.get('timestamp'),
            request.args.get('nonce'), request.get_data(as_text=True))
        import xml.etree.ElementTree as ET
        root = ET.fromstring(plain)
        if root.findtext('MsgType') == 'file':
            media_id = root.findtext('MediaId')
            chat_id = root.findtext('ChatId')
            if media_id and chat_id:
                threading.Thread(target=_wecom_process_file, args=(media_id, chat_id), daemon=True).start()
    except Exception as e:
        print('[wecom] 解析失败:', e)
    return crypto.encrypt_message('success', request.args.get('nonce', ''))

@app.route('/<path:filename>')
def wecom_verify_file(filename):
    """企业微信域名所有权验证：在后台填的校验文件名与内容通过环境变量配置。"""
    cfg_fn = os.environ.get('WECOM_VERIFY_FILENAME')
    if cfg_fn and filename == cfg_fn:
        from flask import Response
        return Response(os.environ.get('WECOM_VERIFY_CONTENT', ''), mimetype='text/plain')
    return 'not found', 404

def _find_free_port(preferred=5000):
    """找到一个可用端口，避免与其他程序冲突"""
    for p in range(preferred, preferred + 100):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            if s.connect_ex(('127.0.0.1', p)) != 0:
                return p
    return preferred

if __name__ == '__main__':
    import threading
    host = '127.0.0.1'
    port = _find_free_port(int(os.environ.get('PORT', 5000)))
    url = f'http://{host}:{port}/'
    # 启动后自动打开默认浏览器
    threading.Timer(1.5, lambda: webbrowser.open(url)).start()
    print('=' * 52)
    print('   电商审单核对工具 已启动')
    print(f'   请在浏览器中访问：{url}')
    print('   本程序仅在本机运行，关闭此窗口即停止服务')
    print('=' * 52)
    app.run(host=host, port=port, debug=False)
